import os
from natsort import natsorted
from openpyxl import Workbook
from openpyxl.styles import PatternFill

dir_path = input("폴더 경로를 입력하세요\n=> ")
excel_path = input("엑셀 경로를 입력하세요\n=> ")

wb = Workbook()
ws = wb.active

headers = ['연번', '파일명', '확장자', '경로']
header_color = PatternFill(start_color='4f81bd',
                           end_color='4f81bd', fill_type='solid')
for col_idx, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col_idx, value=header)
    ws.cell(row=1, column=col_idx).fill = header_color

max_row = ws.max_row + 1
for root, _, files in os.walk(dir_path):
    for idx, file in enumerate(natsorted(files)):
        filename, extension = os.path.splitext(os.path.basename(file))
        extension_wo_dot = extension.lstrip('.')
        ws.cell(row=max_row, column=1, value=idx + 1)
        ws.cell(row=max_row, column=2, value=filename)
        ws.cell(row=max_row, column=3, value=extension_wo_dot)
        ws.cell(row=max_row, column=4, value=os.path.join(root, file))
        max_row += 1

wb.save(excel_path)
